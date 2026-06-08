import streamlit as st
import h5py
import pandas as pd
import numpy as np
from io import BytesIO
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column schema ────────────────────────────────────────────────────────────

MASTER_COLUMNS = [
    # Identity
    "Date Performed",
    "Species / System",
    "Specimen ID",
    "Experimental Condition",
    # Morphology
    "Length",
    "Cross-Sectional Area",
    "Second Moment of Area (I)",
    # Forces & Mechanics
    "Force",
    "Torque",
    "Stress",
    "Muscle Force",
    # Kinematics
    "Strain",
    "Strain Rate",
    "Curvature",
    "Angular Displacement",
    # Stiffness & Damping
    "Elastic Modulus (E)",
    "Flexural Stiffness (EI)",
    "Angular Stiffness",
    "Damping",
    # Frequency Domain
    "Natural Frequency",
    "Resonance Frequency",
    "Frequency Response",
    "Transfer Function",
    # Activation & Timing
    "Activation Timing",
    "Phase",
    "Duty Cycle",
    # Muscle Mechanics
    "Length–Tension Data",
    "Force–Velocity Data",
    "Work / Power Output",
    "Efficiency / Energetic Cost",
    # Advanced Mechanics
    "Passive vs Active Stiffness",
    "Local Stiffness",
    "Local Damping",
    "Viscoelastic Decomposition",
    "Nonlinearity Metrics",
]

# Key aliases: map HDF5 dataset/attribute names → master column names
ALIASES = {
    # ── Date Performed ───────────────────────────────────────────────────────
    "date": "Date Performed",
    "date_performed": "Date Performed",
    "datetime": "Date Performed",
    "start_time_iso": "Date Performed",
    "timestamp": "Date Performed",
    "acquisition_start": "Date Performed",
    "experiment_date": "Date Performed",
    "exp_date": "Date Performed",

    # ── Species / System ─────────────────────────────────────────────────────
    "species": "Species / System",
    "genus_species": "Species / System",
    "organism": "Species / System",
    "system": "Species / System",
    "species_system": "Species / System",
    # your file contains "Micropterus nigricans" (bass) — genus + species fields
    "genus": "Species / System",

    # ── Specimen ID ──────────────────────────────────────────────────────────
    "specimen_id": "Specimen ID",
    "specimen": "Specimen ID",
    "sample_id": "Specimen ID",
    "id": "Specimen ID",

    # ── Experimental Condition ───────────────────────────────────────────────
    "condition": "Experimental Condition",
    "prep_condition": "Experimental Condition",
    "experimental_condition": "Experimental Condition",
    "test_type": "Experimental Condition",         # e.g. "isometric"
    "isometric_mode": "Experimental Condition",
    "simulation_mode": "Experimental Condition",
    "protocol_metadata": "Experimental Condition",
    "config_name": "Experimental Condition",
    "block_sequence": "Experimental Condition",
    "treatment": "Experimental Condition",

    # ── Length ───────────────────────────────────────────────────────────────
    "length": "Length",
    "fishlen": "Length",                            # fishlen_ in your file
    "fishlen_": "Length",
    "test_segment_length_mm": "Length",
    "test_segment_position_mm": "Length",
    "xsec_height": "Length",
    "xsec_width": "Length",

    # ── Cross-Sectional Area ─────────────────────────────────────────────────
    "cross_sectional_area": "Cross-Sectional Area",
    "csa": "Cross-Sectional Area",
    "area": "Cross-Sectional Area",
    "specimen_geometry_heights_mm": "Cross-Sectional Area",
    "specimen_geometry_depths_mm": "Cross-Sectional Area",

    # ── Second Moment of Area (I) ────────────────────────────────────────────
    "second_moment_of_area": "Second Moment of Area (I)",
    "moment_of_area": "Second Moment of Area (I)",
    "specimen_moi_specimen": "Second Moment of Area (I)",
    "i_total_system": "Second Moment of Area (I)",

    # ── Force ────────────────────────────────────────────────────────────────
    "force": "Force",
    "forcetorque": "Force",
    "forcetorque_corrected": "Force",
    "forcetorque_raw": "Force",
    "mean_xforce_stim": "Force",

    # ── Torque ───────────────────────────────────────────────────────────────
    "torque": "Torque",
    "primary_torque_corrected": "Torque",
    "primary_torque_raw": "Torque",
    "inertial_torque_specimen_primary": "Torque",
    "inertial_torque_system_primary": "Torque",
    "inertial_torque_total_primary": "Torque",

    # ── Stress ───────────────────────────────────────────────────────────────
    "stress": "Stress",

    # ── Muscle Force ─────────────────────────────────────────────────────────
    "muscle_force": "Muscle Force",
    "recruitment": "Muscle Force",
    "stim_state": "Muscle Force",

    # ── Strain ───────────────────────────────────────────────────────────────
    "strain": "Strain",
    "strain_pct": "Strain",

    # ── Strain Rate ──────────────────────────────────────────────────────────
    "strain_rate": "Strain Rate",
    "amp_step_vel": "Strain Rate",
    "velocity_exponent": "Strain Rate",

    # ── Curvature ────────────────────────────────────────────────────────────
    "curvature": "Curvature",
    "curvature_1_per_m": "Curvature",

    # ── Angular Displacement ─────────────────────────────────────────────────
    "angular_displacement": "Angular Displacement",
    "angle_measured": "Angular Displacement",
    "angle_cmd": "Angular Displacement",
    "anglevel_cmd": "Angular Displacement",
    "target_deg": "Angular Displacement",
    "max_commanded_rotation_deg": "Angular Displacement",
    "ramp_from_deg": "Angular Displacement",

    # ── Elastic Modulus (E) ──────────────────────────────────────────────────
    "elastic_modulus": "Elastic Modulus (E)",
    "modulus": "Elastic Modulus (E)",
    "young_modulus": "Elastic Modulus (E)",
    "simulation_material": "Elastic Modulus (E)",

    # ── Flexural Stiffness (EI) ──────────────────────────────────────────────
    "flexural_stiffness": "Flexural Stiffness (EI)",
    "ei": "Flexural Stiffness (EI)",
    "force_length_results": "Flexural Stiffness (EI)",

    # ── Angular Stiffness ────────────────────────────────────────────────────
    "angular_stiffness": "Angular Stiffness",

    # ── Damping ──────────────────────────────────────────────────────────────
    "damping": "Damping",

    # ── Natural Frequency ────────────────────────────────────────────────────
    "natural_frequency": "Natural Frequency",
    "daq_ao_do_sample_rate_hz": "Natural Frequency",

    # ── Resonance Frequency ──────────────────────────────────────────────────
    "resonance_frequency": "Resonance Frequency",
    "sono_internal_rate": "Resonance Frequency",

    # ── Frequency Response ───────────────────────────────────────────────────
    "frequency_response": "Frequency Response",

    # ── Transfer Function ────────────────────────────────────────────────────
    "transfer_function": "Transfer Function",

    # ── Activation Timing ────────────────────────────────────────────────────
    "activation_timing": "Activation Timing",
    "activation": "Activation Timing",
    "stim_onset_s": "Activation Timing",
    "stim_t0": "Activation Timing",
    "stim_t1": "Activation Timing",
    "stim_duration_s": "Activation Timing",
    "t_active_start": "Activation Timing",
    "t_active_end": "Activation Timing",
    "prestim_time": "Activation Timing",
    "poststim_time": "Activation Timing",

    # ── Phase ────────────────────────────────────────────────────────────────
    "phase": "Phase",
    "stim_side": "Phase",
    "bilateral_sequential_left_frac": "Phase",
    "prepoststim_sep": "Phase",
    "prepoststim_dur": "Phase",

    # ── Duty Cycle ───────────────────────────────────────────────────────────
    "duty_cycle": "Duty Cycle",
    "pulse_width_ms": "Duty Cycle",
    "stim_pulse_rate": "Duty Cycle",

    # ── Length–Tension Data ──────────────────────────────────────────────────
    "length_tension": "Length–Tension Data",
    "isometric_final": "Length–Tension Data",
    "isometric_initial": "Length–Tension Data",
    "isometric_num_steps": "Length–Tension Data",
    "isometric_stim_params": "Length–Tension Data",

    # ── Force–Velocity Data ──────────────────────────────────────────────────
    "force_velocity": "Force–Velocity Data",

    # ── Work / Power Output ──────────────────────────────────────────────────
    "work": "Work / Power Output",
    "power": "Work / Power Output",
    "power_output": "Work / Power Output",

    # ── Efficiency / Energetic Cost ──────────────────────────────────────────
    "efficiency": "Efficiency / Energetic Cost",
    "energetic_cost": "Efficiency / Energetic Cost",

    # ── Passive vs Active Stiffness ──────────────────────────────────────────
    "passive_stiffness": "Passive vs Active Stiffness",
    "active_stiffness": "Passive vs Active Stiffness",
    "passive_vs_active_stiffness": "Passive vs Active Stiffness",

    # ── Local Stiffness ──────────────────────────────────────────────────────
    "local_stiffness": "Local Stiffness",
    "specimen_geometry_positions_mm": "Local Stiffness",

    # ── Local Damping ────────────────────────────────────────────────────────
    "local_damping": "Local Damping",

    # ── Viscoelastic Decomposition ───────────────────────────────────────────
    "viscoelastic": "Viscoelastic Decomposition",
    "viscoelastic_decomposition": "Viscoelastic Decomposition",

    # ── Nonlinearity Metrics ─────────────────────────────────────────────────
    "nonlinearity": "Nonlinearity Metrics",
    "nonlinearity_metrics": "Nonlinearity Metrics",
}


def normalize_key(key: str) -> str:
    """Lowercase, strip, replace spaces/dashes with underscores."""
    return re.sub(r"[\s\-]+", "_", key.strip().lower())


def summarize_value(val):
    """Convert HDF5 dataset values to a concise string for the cell."""
    if isinstance(val, (np.ndarray,)):
        if val.size == 0:
            return None
        if val.size == 1:
            return float(val.flat[0])
        try:
            return f"array{list(val.shape)}: [{val.flat[0]:.4g} … {val.flat[-1]:.4g}]"
        except Exception:
            return f"array{list(val.shape)}"
    if isinstance(val, bytes):
        return val.decode("utf-8", errors="replace")
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.integer,)):
        return int(val)
    return val


def extract_from_h5(file_bytes: bytes, filename: str) -> dict:
    """
    Walk the H5 file collecting attributes and scalar datasets,
    map them to master columns via ALIASES.
    Returns a dict {master_column: value}.
    """
    row = {col: None for col in MASTER_COLUMNS}

    with h5py.File(BytesIO(file_bytes), "r") as f:

        def collect(name, obj):
            # Collect attributes on every object
            for attr_key, attr_val in obj.attrs.items():
                nk = normalize_key(attr_key)
                master = ALIASES.get(nk)
                if master and row[master] is None:
                    row[master] = summarize_value(attr_val)

            # Collect datasets
            if isinstance(obj, h5py.Dataset):
                nk = normalize_key(name.split("/")[-1])
                master = ALIASES.get(nk)
                if master and row[master] is None:
                    try:
                        row[master] = summarize_value(obj[()])
                    except Exception:
                        pass

        f.visititems(collect)

        # Also check root-level attributes
        for attr_key, attr_val in f.attrs.items():
            nk = normalize_key(attr_key)
            master = ALIASES.get(nk)
            if master and row[master] is None:
                row[master] = summarize_value(attr_val)

    # Tag source file
    row["_source_file"] = filename
    return row


def build_excel(df: pd.DataFrame) -> bytes:
    """Create a styled master Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Data"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    subheader_fill = PatternFill("solid", start_color="2E75B6")
    subheader_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    null_fill = PatternFill("solid", start_color="FFF2CC")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Column groups ───────────────────────────────────────────────────────
    groups = {
        "Identity": ["Date Performed", "Species / System", "Specimen ID", "Experimental Condition"],
        "Morphology": ["Length", "Cross-Sectional Area", "Second Moment of Area (I)"],
        "Forces & Mechanics": ["Force", "Torque", "Stress", "Muscle Force"],
        "Kinematics": ["Strain", "Strain Rate", "Curvature", "Angular Displacement"],
        "Stiffness & Damping": ["Elastic Modulus (E)", "Flexural Stiffness (EI)", "Angular Stiffness", "Damping"],
        "Frequency Domain": ["Natural Frequency", "Resonance Frequency", "Frequency Response", "Transfer Function"],
        "Activation & Timing": ["Activation Timing", "Phase", "Duty Cycle"],
        "Muscle Mechanics": ["Length–Tension Data", "Force–Velocity Data", "Work / Power Output", "Efficiency / Energetic Cost"],
        "Advanced Mechanics": ["Passive vs Active Stiffness", "Local Stiffness", "Local Damping", "Viscoelastic Decomposition", "Nonlinearity Metrics"],
    }

    # Build ordered column list including source file
    ordered_cols = ["Source File"] + MASTER_COLUMNS

    # ── Row 1: Group headers ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.cell(1, 1, "Source File").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = center
    ws.cell(1, 1).border = border

    col_idx = 2  # start after Source File
    for group_name, cols in groups.items():
        start = col_idx
        end = col_idx + len(cols) - 1
        for c in range(start, end + 1):
            cell = ws.cell(1, c)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = center
            cell.border = border
        ws.cell(1, start, group_name)
        if start != end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        col_idx += len(cols)

    # ── Row 2: Column headers ────────────────────────────────────────────────
    ws.row_dimensions[2].height = 36
    for c_idx, col_name in enumerate(ordered_cols, start=1):
        cell = ws.cell(2, c_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # ── Data rows ────────────────────────────────────────────────────────────
    for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
        ws.row_dimensions[r_idx].height = 18
        # Source file col
        cell = ws.cell(r_idx, 1, row.get("_source_file", ""))
        cell.font = Font(name="Arial", size=10, italic=True, color="595959")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        for c_idx, col in enumerate(MASTER_COLUMNS, start=2):
            val = row.get(col)
            cell = ws.cell(r_idx, c_idx)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if val is None or (isinstance(val, float) and np.isnan(val)):
                cell.value = "null"
                cell.fill = null_fill
                cell.font = Font(name="Arial", size=10, color="999999", italic=True)
            else:
                cell.value = val

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28  # source file
    for c_idx, col in enumerate(MASTER_COLUMNS, start=2):
        letter = get_column_letter(c_idx)
        ws.column_dimensions[letter].width = max(18, min(len(col) + 4, 32))

    # Freeze panes below header rows
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="H5 → Master Excel Pipeline",
    page_icon="🧬",
    layout="wide",
)

st.title("🧬 H5 Data Standardization Pipeline")
st.markdown(
    "Upload one or more `.h5` files from any year. The pipeline extracts "
    "standardized fields and compiles them into a single master Excel file. "
    "Missing values are filled with **null**."
)

uploaded_files = st.file_uploader(
    "Upload H5 files",
    type=["h5", "hdf5"],
    accept_multiple_files=True,
)

def dump_h5_structure(file_bytes: bytes) -> list[dict]:
    """Return a flat list of every key, type, shape, and sample value in an H5 file."""
    records = []
    with h5py.File(BytesIO(file_bytes), "r") as f:
        # Root attributes
        for k, v in f.attrs.items():
            records.append({
                "Path": f"/ (attr) → {k}",
                "Type": "attribute",
                "Shape": str(np.array(v).shape) if hasattr(v, "__len__") else "scalar",
                "Sample Value": str(v)[:120],
                "Normalized Key": normalize_key(k),
                "Mapped To": ALIASES.get(normalize_key(k), "—"),
            })

        def visitor(name, obj):
            # Attributes on this object
            for k, v in obj.attrs.items():
                nk = normalize_key(k)
                records.append({
                    "Path": f"{name} (attr) → {k}",
                    "Type": "attribute",
                    "Shape": str(np.array(v).shape) if hasattr(v, "__len__") else "scalar",
                    "Sample Value": str(v)[:120],
                    "Normalized Key": nk,
                    "Mapped To": ALIASES.get(nk, "—"),
                })
            # Dataset itself
            if isinstance(obj, h5py.Dataset):
                nk = normalize_key(name.split("/")[-1])
                try:
                    val = obj[()]
                    if isinstance(val, np.ndarray) and val.size == 0:
                        sample = "<empty array>"
                    elif hasattr(val, "flat"):
                        sample = str(val.flat[0])
                    else:
                        sample = str(val)
                except Exception:
                    sample = "<unreadable>"
                records.append({
                    "Path": name,
                    "Type": "dataset",
                    "Shape": str(obj.shape),
                    "Sample Value": sample[:120],
                    "Normalized Key": nk,
                    "Mapped To": ALIASES.get(nk, "—"),
                })

        f.visititems(visitor)
    return records


if uploaded_files:
    st.markdown(f"**{len(uploaded_files)} file(s) uploaded.** Processing…")

    # ── Debug: H5 Structure Inspector ────────────────────────────────────────
    with st.expander("🔍 Inspect H5 file structure (use this if fields are missing)", expanded=False):
        inspect_file = st.selectbox("Select file to inspect", [uf.name for uf in uploaded_files])
        for uf in uploaded_files:
            if uf.name == inspect_file:
                raw = uf.read()
                uf.seek(0)  # reset so extract_from_h5 can read it again
                try:
                    records = dump_h5_structure(raw)
                    struct_df = pd.DataFrame(records)
                    unmapped = struct_df[struct_df["Mapped To"] == "—"]
                    mapped = struct_df[struct_df["Mapped To"] != "—"]
                    st.markdown(f"**{len(records)} total keys** — ✅ {len(mapped)} mapped, ❌ {len(unmapped)} unmapped")
                    st.markdown("**All keys found in this file:**")
                    st.dataframe(struct_df, use_container_width=True)
                    if not unmapped.empty:
                        st.markdown("**❌ Unmapped keys** (copy these names and share them so the alias table can be updated):")
                        st.dataframe(unmapped[["Path", "Normalized Key", "Sample Value"]], use_container_width=True)
                except Exception as e:
                    st.error(f"Could not read structure: {e}")
                break

    rows = []
    errors = []

    progress = st.progress(0)
    status = st.empty()

    for i, uf in enumerate(uploaded_files):
        status.text(f"Processing: {uf.name}")
        try:
            uf.seek(0)
            row = extract_from_h5(uf.read(), uf.name)
            rows.append(row)
        except Exception as e:
            errors.append((uf.name, str(e)))
        progress.progress((i + 1) / len(uploaded_files))

    status.empty()
    progress.empty()

    if errors:
        with st.expander("⚠️ Files with errors", expanded=True):
            for fname, err in errors:
                st.error(f"**{fname}**: {err}")

    if rows:
        df = pd.DataFrame(rows)

        # ── Preview ──────────────────────────────────────────────────────────
        st.subheader("📋 Preview")
        preview_cols = ["_source_file"] + MASTER_COLUMNS[:8]
        preview_df = df[preview_cols].copy()
        preview_df.columns = ["Source File"] + MASTER_COLUMNS[:8]
        preview_df = preview_df.fillna("null")
        st.dataframe(preview_df, use_container_width=True)

        # ── Coverage stats ───────────────────────────────────────────────────
        st.subheader("📊 Field Coverage")
        coverage = {}
        for col in MASTER_COLUMNS:
            filled = df[col].notna().sum()
            coverage[col] = round(100 * filled / len(df), 1)
        cov_df = pd.DataFrame.from_dict(coverage, orient="index", columns=["% Filled"])
        cov_df["% Filled"] = cov_df["% Filled"].astype(float)

        col1, col2 = st.columns([2, 1])
        with col1:
            st.bar_chart(cov_df)
        with col2:
            st.dataframe(cov_df.style.format("{:.1f}%"), use_container_width=True)

        # ── Download ─────────────────────────────────────────────────────────
        st.subheader("⬇️ Download Master Excel")
        excel_bytes = build_excel(df)
        st.download_button(
            label="📥 Download master_data.xlsx",
            data=excel_bytes,
            file_name="master_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        st.success(
            f"✅ {len(rows)} file(s) processed successfully. "
            f"{sum(1 for e in errors)} error(s)."
        )
    else:
        st.warning("No data could be extracted. Check the error log above.")
else:
    st.info("👆 Upload your H5 files above to get started.")

    with st.expander("ℹ️ How field matching works"):
        st.markdown(
            """
The pipeline scans every **dataset** and **attribute** inside each H5 file and 
matches them to master columns using a flexible alias dictionary.

For example, all of these keys map to **"Force"**:
- `force`, `Force`, `FORCE`

And these map to **"Cross-Sectional Area"**:
- `cross_sectional_area`, `csa`, `area`

If a field isn't found, the cell is marked **null** (highlighted in yellow in Excel).

**Tip**: If your files use custom naming, you can extend the alias table in `app.py` 
under the `ALIASES` dictionary.
            """
        )
